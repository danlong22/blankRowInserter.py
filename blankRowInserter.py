#!python3
#blankRowInserter.py - inserts blank rows in a spreadsheet and saves the new spreadsheet (with blank rows)
#					   to 'page2'
# Usage - The program will ask for the name of the spreadsheet, which should be in the same directory as the program.
#         It will then ask where to put the rows and how many blank rows. The results will be written to 'Sheet2'
import openpyxl
from openpyxl import load_workbook

#change to user specified values
file_name = input('What is the name of the spreadsheet you wish to add blank rows to? \n')
row_number = input('After which row would you like to insert the blank rows? \n')
number_of_rows = input('How many blank rows would you like to insert? \n')

#changes user input to integers
row_number = int(row_number)
number_of_rows = int(number_of_rows)

wb = openpyxl.load_workbook(file_name)

#create another sheet to copy to
if 'Sheet2' not in wb.get_sheet_names():
	wb.create_sheet('Sheet2')

sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

#copies cells before blank lines
for i in range (1, row_number):
	for j in range (sheet1.max_column):
		sheet2.cell(row = i, column = j+1).value = sheet1.cell(row =i, column = j+1).value
#inserts blank lines
for i in range (row_number, row_number+number_of_rows):
	for j in range(sheet1.max_column):
		sheet2.cell(row=i+1, column = j+1).value = ''
#prints everything after the blank rows
for i in range (row_number, sheet1.max_row):
	for j in range(sheet1.max_column):
		sheet2.cell(row = i+1+number_of_rows, column =j+1).value = sheet1.cell(row = i+1, column =j+1).value
wb.save('practiceSpreadsheet.xlsx')